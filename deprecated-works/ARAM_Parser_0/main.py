from sre_parse import expand_template
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import xlsxwriter
import os
from os.path import exists
from selenium.webdriver.chrome.service import Service

time_r = int(input())
LIST_OF_REGIONS = [
    "BR",
    "EUNE",
    "EUW",
    "JP",
    "KR",
    "LAN",
    "LAS",
    "NA",
    "OCE",
    "TR",
    "RU"
]

def get_source_html(url):
    #driver = webdriver.Chrome(
    #    executable_path="chromeDriver\chromedriver.exe"
    #)

    options = webdriver.ChromeOptions()
    ser = Service("C:/Users/DragonettE/.wdm/drivers/chromedriver/win32/101.0.4951.41/chromedriver.exe")
    driver = webdriver.Chrome(service=ser, options=options)
    
    try:
        driver.get(url=url)
        time.sleep(5)
        
        while True:
            if driver.find_elements(by=By.CLASS_NAME, value="livegames"):
                with open("source-html.html", "w", encoding="utf8") as file:
                    file.write(driver.page_source)
                break
            else:
                continue

    except Exception as _ex:
        print(_ex)

    finally:
        driver.close()
        driver.quit()


def get_matches(html_file_path, xlsx_file_path):
    with open(html_file_path, encoding="utf8") as file:
        src = file.read()

    soup = BeautifulSoup(src, 'lxml')
    
    workbook = load_workbook(xlsx_file_path)

    for region in LIST_OF_REGIONS:
       worksheet = workbook["GLOBAL"]
       region_matches = soup.find_all("div", class_=f"cf r100 region-{region}")
       row = worksheet.max_row

       for match in region_matches:
           tables = match.find_all("table")
           left_team, rigth_team = tables[0].find("tbody"), tables[1].find("tbody")
           players_of_left_team, players_of_right_team = left_team.find_all("tr"), rigth_team.find_all("tr")
           
           ch = [] # First five is left team, second five is right team
           nicknames = [] # Analogously
           for player_of_left_team in players_of_left_team:
               lt_character = player_of_left_team.find_all("td")[0].find("a", href=True)['href'].split('/')[-1]
               lt_nicknames = player_of_left_team.find_all("td")[1].find("a").text
               ch.append(lt_character)
               nicknames.append(lt_nicknames)
            

           scha = str(ch[0] + " - " + ch[1] + " - " + ch[2] + " - " + ch[3] + " - " + ch[4])
           worksheet.cell(row=row, column=1).value = scha
           worksheet.cell(row=row, column=2).value = nicknames[0]
           row +=1

    print("Data collected !")
           


    workbook.save("Matches.xlsx")
    workbook.close()

def create_excel():
    print("Creating a Excel file..")
    workbook = xlsxwriter.Workbook('Matches.xlsx')
    worksheet = workbook.add_worksheet("GLOBAL")
    workbook.close()
    print("Excel file is created.")

def main():
    url = "https://lolprofile.net/livegames"
    if not exists("Matches.xlsx"):
        create_excel()
    else:
        print("Matches.xlsx already exists!")

    while True:

        get_source_html(url=url)
        get_matches(html_file_path="source-html.html", xlsx_file_path="Matches.xlsx")
        time.sleep(time_r)


if __name__ == '__main__':
    main()