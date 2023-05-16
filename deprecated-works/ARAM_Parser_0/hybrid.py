from sre_parse import expand_template
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import xlsxwriter
import os
from os.path import exists
from selenium.webdriver.chrome.service import Service

xlsx_file_path = 'Matches.xlsx'

url1 = "https://lolprofile.net/livegames/"
url2 = "https://lolspectator.tv/"
xpath_1 = "/html/body/main/div"
xpath_2 = "/html/body/div[1]/div[3]/div[4]/div[1]/div/div"
LIST_OF_REGIONS = ["BR", "EUNE", "EUW", "JP", "KR", "LAN", "LAS", "NA", "OCE", "TR", "RU"]

def get_page_soup(url, xpath_):

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless")
    ser = Service("C:/Users/DragonettE/.wdm/drivers/chromedriver/win32/101.0.4951.41/chromedriver.exe")
    driver = webdriver.Chrome(service=ser, options=options)
    # Цикл для ожидания прогрузки страницы
    
    try:
        driver.get(url=url)
        time.sleep(3)
        
        while True:
            if driver.find_elements(by=By.XPATH, value=xpath_):
                soup = BeautifulSoup(driver.page_source, 'lxml')
                return soup
            else:
                continue
        driver.quit()
        driver.close()

    except Exception as _ex:
        print(_ex)

    return soup

def get_matches_lp(page_soup):

    for region in LIST_OF_REGIONS:

        region_matches = page_soup.find_all("div", class_=f"cf r100 region-{region}")

        for match in region_matches:
            tables = match.find_all("table")
            left_team = tables[0].find("tbody")
            players_of_left_team = left_team.find_all("tr")
            nickname = players_of_left_team[0].find_all("td")[1].find("a").text

            ch = []

            for player_of_left_team in players_of_left_team:
                lt_character = player_of_left_team.find_all("td")[0].find("a", href=True)['href'].split('/')[-1]
                ch.append(lt_character)

            ch = sorted(ch)
            ch = ' - '.join([str(item) for item in ch])
            parse_into_excel(ch, nickname)


def get_matches_ls(page_soup):

    region_matches = page_soup.find_all("div", class_="flex -mx-2")

    for match in region_matches:
        tables = match.find_all("div", class_="w-1/2 p-1")
        left_team = tables[0]
        players_of_left_team = tables[0].find_all("div", class_="items-center justify-between flex-wrap flex")
        nick = players_of_left_team[0].find("div", class_="block flex-grow flex w-auto").find("span").text
        team = []

        for player_of_left_team in players_of_left_team:
            lt_character = player_of_left_team.find("div", class_="flex flex-shrink-0 w-auto justify-start mr-2"
                                                    ).find("img", alt=True)['alt']
            team.append(lt_character)

        team = sorted(team)
        team = ' - '.join(str(item) for item in team)
        parse_into_excel(team, nick)


def parse_into_excel(team, nick):

    try:
        workbook = load_workbook(xlsx_file_path)
        worksheet = workbook["GLOBAL"]
        row = worksheet.max_row

        worksheet.cell(row=row, column=1).value = team + " ----- " + f"[{nick}]"
        worksheet.cell(row=row + 1, column=1).value = "-"
        row += 1

        workbook.save("Matches.xlsx")
        workbook.close()
    except PermissionError:
        print("Write error: Please close XLSX to complete parsing")
        time.sleep(5)
        parse_into_excel(team, nick)

def create_excel():
    print("Creating a Excel file..")
    workbook = xlsxwriter.Workbook('Matches.xlsx')
    worksheet = workbook.add_worksheet("GLOBAL")
    workbook.close()
    print("Excel file is created.")

def main():
    if not exists("Matches.xlsx"):
        create_excel()
    else:
        print("Matches.xlsx already exists!")

    while True:
        page_soup = get_page_soup(url1, xpath_1)
        get_matches_lp(page_soup)
        print("Games from lol profile are collected")
        print("________________________")
        page_soup2 = get_page_soup(url2, xpath_2)
        get_matches_ls(page_soup2)
        print("Games from lol spectator are collected")
        print("________________________")
        print("Next parse after 100s")
        time.sleep(100)
    

if __name__ == '__main__':
    main()