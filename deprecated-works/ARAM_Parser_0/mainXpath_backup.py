from sre_parse import expand_template
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import xlsxwriter
import os
from os.path import exists
from selenium.webdriver.chrome.service import Service



xlsx_file_path = 'Matches.xlsx'
LIST_OF_REGIONS = [
    "BR",
    "EUNE",
    "EUW",
    "JP",
    "KR",
    "LAN",
    "LAS",
    "NA",
    "OCE",
    "TR",
    "RU"
]

def get_page_soup(url):
    #driver = webdriver.Chrome(
    #    executable_path="chromeDriver\chromedriver.exe"
    #)

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless")

    ser = Service("C:/Users/DragonettE/.wdm/drivers/chromedriver/win32/101.0.4951.41/chromedriver.exe")
    # drive_directory = "C:/Users/DragonettE/.wdm/drivers/chromedriver/win32/101.0.4951.41/chromedriver.exe"
    # driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    driver = webdriver.Chrome(service=ser, options=options)
    # Цикл для ожидания прогрузки страницы
    
    try:
        driver.get(url=url)
        time.sleep(5)
        
        while True:
            if driver.find_elements(by=By.XPATH, value="/html/body/main/div"):
                # with open("source-html.html", "w", encoding="utf8") as file:
                #    file.write(driver.page_source)
                soup = BeautifulSoup(driver.page_source, 'lxml')
                return soup
            else:
                continue

    except Exception as _ex:
        print(_ex)

    finally:
        return driver, soup


def get_matches(page_soup):

    
    workbook = load_workbook(xlsx_file_path)
    for region in LIST_OF_REGIONS:
       worksheet = workbook["GLOBAL"]
       region_matches = page_soup.find_all("div", class_=f"cf r100 region-{region}")
       row = worksheet.max_row

       for match in region_matches:
           tables = match.find_all("table")
           left_team, rigth_team = tables[0].find("tbody"), tables[1].find("tbody")
           players_of_left_team, players_of_right_team = left_team.find_all("tr"), rigth_team.find_all("tr")

           ch = dict.fromkeys(["team"])
           ch ["team"] = []
           nicknames = []
           for player_of_left_team in players_of_left_team:
               lt_character = player_of_left_team.find_all("td")[0].find("a", href=True)['href'].split('/')[-1]
               lt_nicknames = player_of_left_team.find_all("td")[1].find("a").text
               ch ["team"].append(lt_character)
               nicknames.append(lt_nicknames)
            
           ch ["team"] = sorted(ch ["team"])
           ch ["team"] = ' - '.join([str(item) for item in ch ["team"]])
           worksheet.cell(row=row, column=1).value = ch ["team"] + " ----- " + f"[{nicknames[0]}]"
           worksheet.cell(row=row + 1, column=1).value = ""
           row +=1

       print(f"Games from {region} are collected.")
    print("Waiting for next parse !")
    print("________________________")
           

    workbook.save("Matches.xlsx")
    workbook.close()

def create_excel():
    print("Creating a Excel file..")
    workbook = xlsxwriter.Workbook('Matches.xlsx')
    worksheet = workbook.add_worksheet("GLOBAL")
    workbook.close()
    print("Excel file is created.")

def main():
    url = "https://lolprofile.net/livegames"
    if not exists("Matches.xlsx"):
        create_excel()
    else:
        print("Matches.xlsx already exists!")
    #print("Cooldown ?")
    #time_r = int(input())
    while True:
        driver, page_soup = get_page_soup(url)
        get_matches(page_soup)
        #time.sleep(time_r)
        #print("Refreshing driver...")
        #driver.refresh()
        #time.sleep(2)
        break

    print("-- END --")

if __name__ == '__main__':
    main()