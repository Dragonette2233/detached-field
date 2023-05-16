from sre_parse import expand_template
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import xlsxwriter
import os
from os.path import exists
from selenium.webdriver.chrome.service import Service



xlsx_file_path = 'Matches.xlsx'
LIST_OF_REGIONS = [
    "br",
    "eune",
    "euw",
    "jp",
    "kr",
    "lan",
    "las",
    "na",
    "oce",
    "tr",
    "ru"
]

def get_page_soup(url):
    #driver = webdriver.Chrome(
    #    executable_path="chromeDriver\chromedriver.exe"
    #)

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless")

    ser = Service("C:/Users/DragonettE/.wdm/drivers/chromedriver/win32/101.0.4951.41/chromedriver.exe")
    # drive_directory = "C:/Users/DragonettE/.wdm/drivers/chromedriver/win32/101.0.4951.41/chromedriver.exe"
    # driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    driver = webdriver.Chrome(service=ser, options=options)
    # Цикл для ожидания прогрузки страницы
    
    try:
        driver.get(url=url)
        time.sleep(5)
        
        while True:
            if driver.find_elements(by=By.XPATH, value="/html/body/div[1]/div[3]/div[2]/div[2]"):
                # with open("source-html.html", "w", encoding="utf8") as file:
                #    file.write(driver.page_source)
                soup = BeautifulSoup(driver.page_source, 'lxml')
                return soup
            else:
                continue

    except Exception as _ex:
        print(_ex)

    return driver, soup
    driver.refresh()


def get_matches(page_soup):

    workbook = load_workbook(xlsx_file_path)
    worksheet = workbook["GLOBAL"]
    region_matches = page_soup.find_all("div", class_="flex -mx-2") # FLEX [5]
    row = worksheet.max_row

    for match in region_matches:
        tables = match.find_all("div", class_="w-1/2 p-1")
        left_team = tables[0]
        players_of_left_team = tables[0].find_all("div", class_="items-center justify-between flex-wrap flex")


        ch = dict.fromkeys(["team"])
        ch["team"] = []
        nicknames = []
        for player_of_left_team in players_of_left_team:
            lt_character = player_of_left_team.find("div", class_="flex flex-shrink-0 w-auto justify-start mr-2"
                                                    ).find("img", alt=True)['alt']
            lt_nicknames = player_of_left_team.find("div", class_="block flex-grow flex w-auto"
                                                    ).find("span").text
            ch["team"].append(lt_character)
            nicknames.append(lt_nicknames)

        ch["team"] = sorted(ch["team"])
        ch["team"] = ' - '.join([str(item) for item in ch["team"]])
        worksheet.cell(row=row, column=1).value = ch["team"] + " ----- " + f"[{nicknames[0]}]"
        worksheet.cell(row=row + 1, column=1).value = ""
        row += 1


    workbook.save("Matches.xlsx")
    workbook.close()

def create_excel():
    print("Creating a Excel file..")
    workbook = xlsxwriter.Workbook('Matches.xlsx')
    worksheet = workbook.add_worksheet("GLOBAL")
    workbook.close()
    print("Excel file is created.")

def main():
    i = 1
    url = "https://lolspectator.tv/featured/"
    if not exists("Matches.xlsx"):
        create_excel()
    else:
        print("Matches.xlsx already exists!")

    while True:
        for region in LIST_OF_REGIONS:

            page_soup = get_page_soup(url + region)
            get_matches(page_soup)

            print(f"Games from lol spectator are collected {i}")
            i += 1



        print("end")
        break


        


if __name__ == '__main__':
    main()